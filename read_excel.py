import openpyxl
import json

wb = openpyxl.load_workbook(r"C:\Users\Faye\Desktop\集运业务流量包数据统计_20260511.xlsx", read_only=True)
print("Sheets:", wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ===")
    rows = list(ws.iter_rows(values_only=True))
    if rows:
        print(f"Total rows: {len(rows)}")
        print(f"Headers (row 1): {rows[0]}")
        print(f"Row 2 sample: {rows[1] if len(rows) > 1 else 'N/A'}")
        print(f"Row 3 sample: {rows[2] if len(rows) > 2 else 'N/A'}")
        print(f"Row 4 sample: {rows[3] if len(rows) > 3 else 'N/A'}")
