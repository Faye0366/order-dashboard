import openpyxl
import json
from collections import defaultdict
from datetime import datetime, timedelta

print("Loading Excel file...")
wb = openpyxl.load_workbook(r"C:\Users\Faye\Desktop\WorkBuddy数据看板\集运业务流量包数据统计.xlsx", read_only=True)
ws = wb["数据源"]

rows = list(ws.iter_rows(values_only=True))
headers = rows[0]

col_map = {h: i for i, h in enumerate(headers)}
date_col = col_map['日期']
success_col = col_map['成功订单']
prod_name_col = col_map['统一产品名称']
channel_col = col_map['投放点位']
type_col = col_map['类型']
baoti_col = col_map['包体']

data_liuliang = []
data_kuandai = []
data_ll_nods = []

for row in rows[1:]:
    d = row[date_col]
    if d is None:
        continue
    if isinstance(d, datetime):
        date = d
    elif isinstance(d, str):
        try:
            date = datetime.strptime(d[:10], "%Y-%m-%d")
        except:
            continue
    else:
        continue

    row_type = row[type_col]
    success = row[success_col]
    if success is None:
        success = 0
    try:
        success = int(success)
    except:
        success = 0

    prod = row[prod_name_col] or '未知'
    channel = row[channel_col] or '未知'
    baoti = str(row[baoti_col]).strip() if row[baoti_col] else ''

    item = {
        'date': date,
        'success': success,
        'prod': str(prod),
        'channel': str(channel),
        'baoti': baoti,
        'month': date.strftime('%Y-%m'),
        'day': date.strftime('%Y-%m-%d'),
        'year': date.strftime('%Y'),
    }

    if str(row_type).strip() == '流量包':
        data_liuliang.append(item)
        if baoti != '短时包':
            data_ll_nods.append(item)
    else:
        data_kuandai.append(item)

print(f"流量包行数: {len(data_liuliang)}, 宽带行数: {len(data_kuandai)}")

def calc_stats(data):
    all_dates = [d['date'] for d in data]
    if not all_dates:
        return {}, {}, {}
    min_date = min(all_dates)
    max_date = max(all_dates)
    annual = defaultdict(int)
    for d in data:
        annual[d['year']] += d['success']
    monthly = defaultdict(int)
    for d in data:
        monthly[d['month']] += d['success']
    daily_by_month = defaultdict(lambda: defaultdict(int))
    for d in data:
        daily_by_month[d['month']][d['day']] += d['success']
    monthly_daily = {}
    for month, days in daily_by_month.items():
        month_data = {}
        for day_str, count in sorted(days.items()):
            month_data[day_str] = count
        monthly_daily[month] = month_data
    return dict(sorted(annual.items())), dict(sorted(monthly.items())), monthly_daily, (min_date, max_date)

annual_ll, monthly_ll, daily_ll, range_ll = calc_stats(data_liuliang)
annual_kd, monthly_kd, daily_kd, range_kd = calc_stats(data_kuandai)
annual_nods, monthly_nods, daily_nods, range_nods = calc_stats(data_ll_nods)

def calc_channel(data):
    result = defaultdict(lambda: defaultdict(int))
    for d in data:
        result[d['month']][d['channel']] += d['success']
    return {k: dict(sorted(v.items(), key=lambda x: -x[1])) for k, v in sorted(result.items())}

def calc_daily_channel(data):
    """逐日分渠道汇总 {month: {day: {channel: count}}}"""
    result = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    for d in data:
        result[d['month']][d['day']][d['channel']] += d['success']
    # 转成普通 dict
    out = {}
    for month, days in sorted(result.items()):
        out[month] = {}
        for day, chs in sorted(days.items()):
            out[month][day] = dict(sorted(chs.items(), key=lambda x: -x[1]))
    return out

def calc_daily_channel_baoti(data):
    """逐日分渠道×包体汇总 {month: {day: {channel: {baoti: count}}}}"""
    result = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(int))))
    for d in data:
        result[d['month']][d['day']][d['channel']][d['baoti']] += d['success']
    # 转成普通 dict
    out = {}
    for month, days in sorted(result.items()):
        out[month] = {}
        for day, chs in sorted(days.items()):
            out[month][day] = {}
            for ch, baotis in sorted(chs.items()):
                out[month][day][ch] = dict(sorted(baotis.items(), key=lambda x: -x[1]))
    return out

daily_ch_ll = calc_daily_channel(data_liuliang)
daily_ch_nods = calc_daily_channel(data_ll_nods)  # 不含短时包的逐日分渠道
daily_ch_baoti_ll = calc_daily_channel_baoti(data_liuliang)  # 逐日分渠道×包体

ch_ll = calc_channel(data_liuliang)
ch_kd = calc_channel(data_kuandai)
ch_nods = calc_channel(data_ll_nods)

def calc_top10(data):
    result = defaultdict(lambda: defaultdict(int))
    for d in data:
        result[d['month']][d['prod']] += d['success']
    top10 = {}
    for month, prods in sorted(result.items()):
        top10[month] = sorted(prods.items(), key=lambda x: -x[1])[:10]
    return top10

top10_ll = calc_top10(data_liuliang)
top10_kd = calc_top10(data_kuandai)
top10_nods = calc_top10(data_ll_nods)

# ---- Process 积分数据 sheet ----
ws_jf = wb["积分数据"]
jf_rows = list(ws_jf.iter_rows(values_only=True))
jf_headers = jf_rows[0]
jf_col_map = {h: i for i, h in enumerate(jf_headers)}
jf_date_col = jf_col_map['日期']
jf_dan_col = jf_col_map['完成单量']
jf_jf_col = jf_col_map['完成积分']
jf_lv_col = jf_col_map['完成率']

monthly_jf = {}
daily_jf = {}

for row in jf_rows[1:]:
    d = row[jf_date_col]
    if d is None:
        continue
    if isinstance(d, datetime):
        date = d
    elif isinstance(d, str):
        try:
            date = datetime.strptime(d[:10], "%Y-%m-%d")
        except:
            continue
    else:
        continue

    month_key = date.strftime('%Y-%m')
    day_key = date.strftime('%Y-%m-%d')
    dan = int(row[jf_dan_col] or 0)
    jf_val = float(row[jf_jf_col] or 0)
    lv_val = float(row[jf_lv_col] or 0)

    if dan == 0 and jf_val == 0:
        continue

    if month_key not in daily_jf:
        daily_jf[month_key] = {}
        monthly_jf[month_key] = {"完成积分": 0, "完成单量": 0}

    daily_jf[month_key][day_key] = {
        "完成积分": round(jf_val, 1),
        "完成单量": dan,
        "完成率": round(lv_val, 4)
    }
    monthly_jf[month_key]["完成积分"] += jf_val
    monthly_jf[month_key]["完成单量"] += dan

for mk in monthly_jf:
    monthly_jf[mk]["完成积分"] = round(monthly_jf[mk]["完成积分"], 1)
    monthly_jf[mk]["完成率"] = round(monthly_jf[mk]["完成积分"] / 250000, 4)

print(f"积分有效月数: {sorted(monthly_jf.keys())}")
print(f"5月完成积分: {monthly_jf.get('2026-05', {})}")

result = {
    'annual': annual_ll,
    'monthly': monthly_ll,
    'daily_by_month': daily_ll,
    'daily_by_month_channel': daily_ch_ll,
    'daily_by_month_channel_nods': daily_ch_nods,  # 不含短时包的逐日分渠道
    'daily_by_month_channel_baoti': daily_ch_baoti_ll,  # 逐日分渠道×包体
    'monthly_channel': ch_ll,
    'monthly_top10': top10_ll,
    'annual_nods': annual_nods,
    'monthly_nods': monthly_nods,
    'daily_by_month_nods': daily_nods,
    'monthly_channel_nods': ch_nods,
    'monthly_top10_nods': top10_nods,
    'annual_kuandai': annual_kd,
    'monthly_kuandai': monthly_kd,
    'daily_by_month_kuandai': daily_kd,
    'monthly_channel_kuandai': ch_kd,
    'monthly_top10_kuandai': top10_kd,
    'monthly_jf': monthly_jf,
    'daily_jf': daily_jf,
    'meta': {
        'liuliang_rows': len(data_liuliang),
        'liuliang_nods_rows': len(data_ll_nods),
        'kuandai_rows': len(data_kuandai),
        'liuliang_range': f"{range_ll[0].strftime('%Y-%m-%d')} to {range_ll[1].strftime('%Y-%m-%d')}",
        'kuandai_range': f"{range_kd[0].strftime('%Y-%m-%d')} to {range_kd[1].strftime('%Y-%m-%d')}" if data_kuandai else 'N/A',
    }
}

with open(r"C:\Users\Faye\WorkBuddy\2026-05-11-task-1\dashboard_data.json", 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

print("Done. dashboard_data.json updated.")
